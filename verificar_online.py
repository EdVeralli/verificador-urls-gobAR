"""
verificar_online.py
-------------------
Lee las URLs del archivo 'urls 2.xlsx' (columna A, sin encabezado),
las limpia, verifica si cada una está online (HTTP < 400) y genera
un nuevo archivo 'resultado_urls2.xlsx' con los resultados coloreados.

Uso:
    python verificar_online.py

Requisitos:
    pip install requests openpyxl
"""

import os
import sys
import concurrent.futures
from urllib.parse import urlparse

# Forzar UTF-8 en la consola de Windows
if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

try:
    import requests
except ImportError:
    print("ERROR: Falta 'requests'. Ejecutá: pip install requests")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
except ImportError:
    print("ERROR: Falta 'openpyxl'. Ejecutá: pip install openpyxl")
    sys.exit(1)

# ──────────────────────────────────────────────
# Configuración
# ──────────────────────────────────────────────
XLSX_INPUT  = "urls 2.xlsx"
XLSX_OUTPUT = "resultado_urls2.xlsx"
MAX_WORKERS = 15
TIMEOUT     = 12

COLOR_ONLINE   = "C6EFCE"
COLOR_OFFLINE  = "FFC7CE"
COLOR_INVALIDA = "D9D9D9"
COLOR_WARNING  = "FFEB9C"
COLOR_HEADER   = "2F5496"

HEADERS_HTTP = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 Chrome/120.0.0.0 Safari/537.36"
    )
}


# ──────────────────────────────────────────────
# Funciones
# ──────────────────────────────────────────────
def limpiar_url(url):
    """Elimina caracteres de puntuación trailing que no forman parte de la URL."""
    return url.rstrip(".,)*")


def es_invalida(url):
    """Devuelve mensaje si la URL tiene formato inválido DESPUÉS de limpiar."""
    if "..." in url:
        return "Contiene puntos suspensivos (...)"
    if "*" in url:
        return "Contiene asterisco (*) en el medio de la URL"
    return None


def verificar_url(item):
    idx, url_orig = item
    url = limpiar_url(url_orig.strip()) if url_orig else ""

    if not url:
        return idx, url_orig, "", "SIN URL", COLOR_INVALIDA

    motivo = es_invalida(url)
    if motivo:
        return idx, url_orig, url, f"INVÁLIDA — {motivo}", COLOR_INVALIDA

    try:
        parsed = urlparse(url)
        if not parsed.scheme or not parsed.netloc:
            return idx, url_orig, url, "FORMATO INVÁLIDO", COLOR_INVALIDA
    except Exception:
        return idx, url_orig, url, "FORMATO INVÁLIDO", COLOR_INVALIDA

    try:
        resp = requests.head(url, timeout=TIMEOUT, allow_redirects=True, headers=HEADERS_HTTP)
        code = resp.status_code
        if code == 405:
            resp = requests.get(url, timeout=TIMEOUT, allow_redirects=True,
                                headers=HEADERS_HTTP, stream=True)
            code = resp.status_code

        if code < 400:
            return idx, url_orig, url, f"✅ ONLINE (HTTP {code})", COLOR_ONLINE
        elif code == 404:
            return idx, url_orig, url, "❌ NO ENCONTRADA (404)", COLOR_OFFLINE
        elif code == 403:
            return idx, url_orig, url, "⚠️ ACCESO DENEGADO (403)", COLOR_WARNING
        else:
            return idx, url_orig, url, f"❌ ERROR HTTP {code}", COLOR_OFFLINE

    except requests.exceptions.SSLError:
        return idx, url_orig, url, "❌ ERROR SSL", COLOR_OFFLINE
    except requests.exceptions.ConnectionError:
        return idx, url_orig, url, "❌ SIN CONEXIÓN", COLOR_OFFLINE
    except requests.exceptions.Timeout:
        return idx, url_orig, url, "❌ TIMEOUT", COLOR_OFFLINE
    except Exception as e:
        return idx, url_orig, url, f"❌ ERROR: {str(e)[:60]}", COLOR_OFFLINE


def leer_urls(path):
    """Lee una sola columna de URLs desde el Excel (sin encabezado)."""
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    items = []
    excluidas = 0
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        val = row[0]
        if not val or not str(val).strip():
            continue
        if "descubrir" in str(val).lower():
            excluidas += 1
            continue
        items.append((row_idx, str(val).strip()))
    return items, excluidas


def guardar_resultado(path, results):
    """Genera un nuevo Excel con: #, URL Original, URL Limpiada, Estado Online."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resultados"

    # Encabezado
    headers = ["#", "URL ORIGINAL", "URL LIMPIADA", "ESTADO ONLINE"]
    ws.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # Datos
    for num, (idx, url_orig, url_limp, estado, color) in enumerate(results, 1):
        ws.append([num, url_orig, url_limp, estado])
        for col_idx in range(1, 5):
            cell = ws.cell(row=num + 1, column=col_idx)
            cell.alignment = Alignment(wrap_text=True)
        estado_cell = ws.cell(row=num + 1, column=4)
        estado_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    # Ancho de columnas
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 70
    ws.column_dimensions["C"].width = 70
    ws.column_dimensions["D"].width = 30

    wb.save(path)


# ──────────────────────────────────────────────
# Main
# ──────────────────────────────────────────────
def main():
    if not os.path.exists(XLSX_INPUT):
        print(f"ERROR: No se encontró '{XLSX_INPUT}'.")
        sys.exit(1)

    print(f"\n{'='*65}")
    print(f"  Verificador Online de URLs — gob.ar")
    print(f"{'='*65}")

    items, excluidas = leer_urls(XLSX_INPUT)
    total = len(items)
    print(f"  Entrada          : {XLSX_INPUT}")
    print(f"  URLs a verificar : {total}")
    print(f"  Excluidas (descubrir): {excluidas}")
    print(f"  Simultáneas      : {MAX_WORKERS}")
    print(f"  Timeout          : {TIMEOUT}s")
    print(f"{'='*65}\n")

    resultados_map = {}
    with concurrent.futures.ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {executor.submit(verificar_url, item): item for item in items}
        done = 0
        for future in concurrent.futures.as_completed(futures):
            idx, url_orig, url_limp, estado, color = future.result()
            resultados_map[idx] = (idx, url_orig, url_limp, estado, color)
            done += 1
            icon = "✅" if "ONLINE" in estado else ("⚠️" if "DENEGADO" in estado or "INVÁLIDA" in estado else "❌")
            print(f"  {icon} [{done:3d}/{total}] {estado:<35} {url_limp[:50]}")

    # Ordenar por índice original
    results = [resultados_map[idx] for idx, _ in items if idx in resultados_map]

    guardar_resultado(XLSX_OUTPUT, results)

    # Resumen
    todos = results
    online   = [r for r in todos if "ONLINE" in r[3]]
    no_enc   = [r for r in todos if "404" in r[3]]
    denegado = [r for r in todos if "403" in r[3]]
    invalida = [r for r in todos if "INVÁLIDA" in r[3] or "FORMATO" in r[3]]
    otros    = [r for r in todos if r not in online + no_enc + denegado + invalida]

    print(f"\n{'='*65}")
    print(f"  RESUMEN FINAL")
    print(f"{'='*65}")
    print(f"  ✅ ONLINE          : {len(online)}")
    print(f"  ❌ NO ENCONTRADAS  : {len(no_enc)}")
    print(f"  ⚠️  ACCESO DENEGADO: {len(denegado)}")
    print(f"  🚫 INVÁLIDAS       : {len(invalida)}")
    print(f"  ⚠️  OTROS ERRORES  : {len(otros)}")
    print(f"  🚫 EXCLUIDAS (descubrir): {excluidas}")
    print(f"  {'─'*45}")
    print(f"  TOTAL              : {total + excluidas}")
    print(f"{'='*65}")
    print(f"\n  Resultado guardado en: {XLSX_OUTPUT}\n")


if __name__ == "__main__":
    main()
